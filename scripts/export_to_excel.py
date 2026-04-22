import pandas as pd
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

PROC_DIR  = Path("data/processed")
RAW_DIR   = Path("data/raw")
EXCEL_DIR = Path("excel")
EXCEL_DIR.mkdir(exist_ok=True)


def style_header(ws, row: int, ncols: int) -> None:
    fill = PatternFill("solid", fgColor="1F3864")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = fill; c.font = font
        c.alignment = Alignment(horizontal="center")


def write_historicals_tab(wb: Workbook, ticker: str) -> None:
    df = pd.read_csv(PROC_DIR / f"{ticker}_historicals.csv")
    ws = wb.create_sheet("Historicals")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{ticker} — Historical Financials (yfinance · Python-Generated)"
    ws["A1"].font = Font(bold=True, size=12, color="1F3864")
    ws.merge_cells("A1:N1")
    ws["A2"] = "Do not edit — regenerate by running export_to_excel.py"
    ws["A2"].font = Font(italic=True, color="888888", size=9)

    display = [
        ("fiscal_year", "Year", None),
        ("revenue", "Revenue", "#,##0"),
        ("gross_profit", "Gross Profit", "#,##0"),
        ("ebit", "EBIT", "#,##0"),
        ("net_income", "Net Income", "#,##0"),
        ("gross_margin_pct", "Gross Margin", "0.0%"),
        ("ebit_margin_pct", "EBIT Margin", "0.0%"),
        ("tax_rate", "Tax Rate", "0.0%"),
        ("depreciation_amortization", "D&A", "#,##0"),
        ("capex", "CapEx", "#,##0"),
        ("free_cash_flow", "Free Cash Flow", "#,##0"),
        ("total_debt", "Total Debt", "#,##0"),
        ("cash", "Cash", "#,##0"),
        ("net_working_capital", "Net WC", "#,##0"),
    ]
    display = [(c,l,f) for c,l,f in display if c in df.columns]

    for i, (_,lbl,_) in enumerate(display, 1):
        ws.cell(row=4, column=i, value=lbl)
    style_header(ws, 4, len(display))

    for r, row_data in df[[c for c,_,_ in display]].iterrows():
        for ci, ((col,_,fmt), val) in enumerate(zip(display, row_data), 1):
            cell = ws.cell(row=r+5, column=ci, value=val)
            if fmt: cell.number_format = fmt

    for col in ws.columns:
        w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w+3, 22)


def write_profile_tab(wb: Workbook, ticker: str) -> None:
    p = RAW_DIR / f"{ticker}_profile.json"
    if not p.exists(): return
    with open(p) as f: info = json.load(f)
    ws = wb.create_sheet("Profile")
    ws.sheet_view.showGridLines = False
    ws["A1"] = f"{ticker} — Company Profile (via yfinance)"
    ws["A1"].font = Font(bold=True, size=12, color="1F3864")
    fields = [
        ("Company", info.get("longName")),
        ("Sector", info.get("sector")),
        ("Beta", info.get("beta")),
        ("Shares Outstanding", info.get("sharesOutstanding")),
        ("Market Cap", info.get("marketCap")),
        ("Current Price", info.get("currentPrice")),
        ("52w High", info.get("fiftyTwoWeekHigh")),
        ("52w Low", info.get("fiftyTwoWeekLow")),
    ]
    for i, (label, value) in enumerate(fields, 3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, color="444444")
        ws.cell(row=i, column=2, value=value)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 28


def main(ticker: str = "MSFT") -> None:
    wb = Workbook()
    wb.remove(wb.active)
    write_historicals_tab(wb, ticker)
    write_profile_tab(wb, ticker)
    for tab in ["Assumptions", "Forecast", "DCF", "Sensitivity", "Summary"]:
        ws = wb.create_sheet(tab)
        ws["A1"] = f"[{tab} — Build manually in Excel]"
        ws["A1"].font = Font(color="888888", italic=True)
    out = EXCEL_DIR / f"{ticker}_DCF_Model.xlsx"
    wb.save(out)
    print(f"✓ Workbook saved → {out}")
    print("  Tabs: Historicals, Profile, Assumptions, Forecast, DCF, Sensitivity, Summary")


if __name__ == "__main__":
    main("MSFT")